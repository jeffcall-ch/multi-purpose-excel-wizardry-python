# This script filters the loads excel file extracted for Presenzano to those supports
# which are posing loads on supplier steel structure (with or without sec steel in supplier scope)
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from rich import print

# import TAV_support lists created by hand:
# 1. secondary steel supplied by Enclosure supplier
# 2. or supports attached directly to primary steel
TAV_support_xls = r"C:\Users\50000700\OneDrive - ansaldoenergiagroup\Desktop\TAV_Enclosure_Loads\TAV Enclosure Supplier Input Supports.xlsx"
df_TAV_PRI_supp = pd.read_excel(TAV_support_xls, sheet_name='PRI_ALL', header=[0])
df_TAV_SEC_supp = pd.read_excel(TAV_support_xls, sheet_name='SEC_ALL', header=[0])
df_TAV_SEC_and_PRI_supp = df_TAV_PRI_supp.append(df_TAV_SEC_supp)
pri_supports = sorted(df_TAV_PRI_supp['Pipe Support KKS'].tolist())
sec_supports = sorted(df_TAV_SEC_supp['Pipe Support KKS'].tolist())

# read loads from Presenzano extract - highest loads so far to be used as standard
loads_xls = r"C:\Users\50000700\OneDrive - ansaldoenergiagroup\Desktop\TAV_Enclosure_Loads\TAV_Support_Loads_from_PRE.xlsx"
df_loads = pd.read_excel(loads_xls, header=[1])
print(df_loads)
print(df_loads.size)
print(df_loads.keys())

# defining loads for NOT stress calculated supports. I used AD2000 water filled thicker pipe values.
# below dict defines {nps:kN} values
not_stress_calc_loads = {0.5: -0.1, 1: -0.1, 1.5: -0.15, 2: -0.2, 3: -0.4, 4: -0.5}

# drop all duplicate KKS. Only one shall be present
df_loads = df_loads.drop_duplicates(subset=['Pipe Support KKS'])

# filter load input to only those KKS which are in scop of supplier (2x lists concat)
df_loads_filtered = df_loads[df_loads['Pipe Support KKS'].isin(sec_supports + pri_supports)]
# merge with df_TAV_SEC_and_PRI_supp in order to show group column
df_loads_filtered = df_loads_filtered.merge(df_TAV_SEC_and_PRI_supp, how='left')
# add "X" to supplier scope col
df_loads_filtered['Sec. steel in supplier scope'] = np.where(df_loads_filtered['Pipe Support KKS'].isin(sec_supports),
                                                             "X", " ")
# group and sort
df_loads_filtered = df_loads_filtered.groupby('Sec. steel in supplier scope').apply(
    lambda x: x.sort_values('Pipe Support KKS'))


# write load value to supports which are not stress calculated
def set_kN(row):
    if (row['NPS Pipe'] in not_stress_calc_loads.keys()) and (row['-Z.1'] == "-"):
        return not_stress_calc_loads[row['NPS Pipe']]
    else:
        return row['-Z.1']


df_loads_filtered['-Z.1'] = df_loads_filtered.apply(lambda row: set_kN(row), axis=1)
df_loads_filtered = df_loads_filtered.sort_values(by=['Group'])

print(df_loads_filtered)
print(df_loads_filtered.size)

df_loads_filtered.to_excel(
    r"C:\Users\50000700\OneDrive - ansaldoenergiagroup\Desktop\TAV_Enclosure_Loads\FINAL_TAV_List_For_Supplier.xlsx")

# Styling the created output excel file
# read the created excel file and give border to grouped rows
wb = openpyxl.load_workbook(
    r"C:\Users\50000700\OneDrive - ansaldoenergiagroup\Desktop\TAV_Enclosure_Loads\FINAL_TAV_List_For_Supplier.xlsx")
ws = wb.active

# helper variables to create upper and lower border of the groups
group_starter_row = 1
group_end_row = 1
# dict of starter and end rows of groups {starter : end}
groups = {}
previous_group = 0
row_counter = 1

# style variables
thin = Side(border_style="thin", color="000000")
border = Border(top=thin, left=thin, right=thin, bottom=thin)
top = Border(top=border.top)
left = Border(left=border.left)
right = Border(right=border.right)
bottom = Border(bottom=border.bottom)

# check out https://openpyxl.readthedocs.io/en/stable/styles.html#styling-merged-cells
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
    current_group = row[22]
    if current_group != previous_group:
        groups[group_starter_row] = group_end_row
        rows = ws[f"A{group_starter_row}:W{group_end_row}"]
        for cell in rows[0]:
            cell.border = cell.border + top
        for cell in rows[-1]:
            cell.border = cell.border + bottom
        group_starter_row = row_counter
        previous_group = current_group
    group_end_row = row_counter
    row_counter += 1

wb.save(r"C:\Users\50000700\OneDrive - ansaldoenergiagroup\Desktop\TAV_Enclosure_Loads\FINAL_TAV_List_For_Supplier_Styled.xlsx")
