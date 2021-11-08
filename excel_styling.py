import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

# read the created excel file and give border to grouped rows
wb = openpyxl.load_workbook(
    r"C:\Users\50000700\OneDrive - ansaldoenergiagroup\Desktop\Navis_xml\TAV_List_For_Supplier.xlsx")
ws = wb.active

# helper variables to create upper and lower border of the groups
group_starter_row = 1
group_end_row = 1
# dict of starter and end rows of groups {starter : end}
groups = {}
previous_group = 0
row_counter = 1

# style variables
# check out https://openpyxl.readthedocs.io/en/stable/styles.html#styling-merged-cells
thin = Side(border_style="thin", color="000000")
border = Border(top=thin, left=thin, right=thin, bottom=thin)
top = Border(top=border.top)
left = Border(left=border.left)
right = Border(right=border.right)
bottom = Border(bottom=border.bottom)

for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
    current_group = row[22]
    if current_group != previous_group:
        groups[group_starter_row] = group_end_row
        rows = ws[f"A{group_starter_row}:W{group_end_row}"]
        for cell in rows[0]:
            cell.border = cell.border + top
        for cell in rows[-1]:
            cell.border = cell.border + bottom

        group_starter_row = row_counter
        previous_group = current_group
    group_end_row = row_counter
    row_counter += 1

wb.save(r"C:\Users\50000700\OneDrive - ansaldoenergiagroup\Desktop\Navis_xml\TAV_List_For_Supplier_Styled.xlsx")
